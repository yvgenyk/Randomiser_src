from PyQt4 import QtGui
from PyQt4.Qt import QListWidgetItem
import time
import sys
import main_design
from openpyxl import load_workbook
from openpyxl import Workbook
from random import randint
import json
import gc


class LetsPlay(QtGui.QMainWindow, main_design.Ui_Form):
    def __init__(self, parent=None):
        super(LetsPlay, self).__init__(parent)
        self.setupUi(self)
        self.loadedFilePath = ""

        self.startBtn.clicked.connect(self.start_p)
        self.loadBtn.clicked.connect(self.open_file)
        self.delBtn.clicked.connect(self.delete_added_strings)


    def start_p(self):

        self.dict = {}
        self.statusLable.setText("")
        time.sleep(1)
        if self.psrLine.text() == "" or self.newFileName.text() == "" or self.loadLable.text() == "":
            self.statusLable.setText("Error!!!")
            QtGui.QMessageBox.question(self, 'Missing input', "Input is missing", QtGui.QMessageBox.Ok)
        else:
            self.psr = int(self.psrLine.text())
            self.new_name = self.newFileName.text()
            wb = load_workbook(filename=self.loadedFilePath)
            self.ws = wb.active
            self.counter = self.num_of_strings()
            self.num_of_random_strings = int(self.counter * (self.psr / 100)) + 1
            self.dict = self.create_spread_list()
            self.random_spread_strings()
            gc.collect()
            self.statusLable.setText("Done!")

    def open_file(self):
        self.loadedFilePath = QtGui.QFileDialog.getOpenFileName(self, 'Open File', "*.xlsx")
        fileName = self.loadedFilePath.split('/')
        self.loadLable.setText('%s' % fileName[len(fileName)-1])

    def delete_added_strings(self):

        if self.loadLable.text() == "":
            QtGui.QMessageBox.question(self, 'Missing input', "Input is missing", QtGui.QMessageBox.Ok)
        else:
            wb = load_workbook(filename=self.loadedFilePath)
            ws = wb.active

            wb_new = Workbook()
            ws_new = wb_new.active
            new_counter = 1

            for i in range(ws.max_row):
                if ws.cell(row=(i+1), column=2).value != "*":
                    ws_new.cell(row=new_counter, column=1).value = ws.cell(row=(i + 1), column=1).value
                    new_counter += 1

            wb_new.save(filename=(self.new_name + "_old.xlsx"))

    def random_spread_strings(self):
        random_nums = []
        for i in range(self.num_of_random_strings):
            random_nums.append(randint(1, self.counter))

        random_nums.sort()

        wb_new = Workbook()
        ws_new = wb_new.active
        counter = 0
        insert_index = 0
        new_row_counter = 1
        all_in = False

        for i in range(self.ws.max_row):

            if counter == random_nums[insert_index] and not all_in:
                for string in self.dict[insert_index]:
                    ws_new.cell(row=new_row_counter, column=1).value = string
                    ws_new.cell(row=new_row_counter, column=2).value = "*"
                    new_row_counter += 1

                if insert_index < (len(random_nums) - 1):
                    insert_index += 1
                else:
                    all_in = True

            if self.ws.cell(row = (i + 1), column = 1).value:
                split = self.ws.cell(row=(i + 1), column=1).value.split("<")
                if split[len(split) - 1] == "/trans-unit>":
                    ws_new.cell(row=new_row_counter, column=1).value = self.ws.cell(row = (i + 1), column = 1).value
                    counter += 1
                    new_row_counter += 1
                else:
                    ws_new.cell(row=new_row_counter, column=1).value = self.ws.cell(row=(i + 1), column=1).value
                    new_row_counter += 1
            else:
                ws_new.cell(row=new_row_counter, column=1).value = self.ws.cell(row=(i + 1), column=1).value
                new_row_counter += 1

        wb_new.save(filename="./New_Files/" + (self.new_name + ".xlsx"))

    def create_spread_list(self):
        counter = 0
        saved_strings = {}

        for i in range(self.ws.max_row - 1):
            if counter >= self.num_of_random_strings:
                break
            else:
                string = []
                if self.ws.cell(row=(i + 1), column=1).value:
                    split = self.ws.cell(row=(i + 1), column=1).value.split("<")
                    if len(split[len(split) - 1]) > 13:
                        if split[len(split) - 1][:13] == "trans-unit id":
                            string.append(self.ws.cell(row=(i + 1), column=1).value)
                            finished = False
                            while not finished:
                                i += 1
                                if self.ws.cell(row=(i + 1), column=1).value:
                                    split = self.ws.cell(row=(i + 1), column=1).value.split("<")
                                    if split[len(split) - 1] == "/trans-unit>":
                                        string.append(self.ws.cell(row=(i + 1), column=1).value)
                                        counter += 1
                                        saved_strings[counter - 1] = string
                                        finished = True
                                    else:
                                        string.append(self.ws.cell(row=(i + 1), column=1).value)
        return saved_strings

    def num_of_strings(self):
        counter = 0
        for i in range(self.ws.max_row - 1):
            if self.ws.cell(row = (i + 1), column = 1).value:
                split = self.ws.cell(row = (i + 1), column = 1).value.split("<")
                if len(split[len(split) - 1]) > 13:
                    if split[len(split) - 1][:13] == "trans-unit id":
                        counter += 1

        return counter

def main():
    app = QtGui.QApplication(sys.argv)
    app.setStyle('cleanlooks')
    form = LetsPlay()
    form.show()
    app.exec_()

if __name__ == '__main__':
    main()